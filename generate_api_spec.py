import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

C_NAVY="1F3864"; C_DARK="2E4057"; C_WHITE="FFFFFF"
C_GET="D9EAD3"; C_GET_FG="38761D"
C_POST="FCE5CD"; C_POST_FG="B45F06"
C_PUT="FFF2CC"; C_PUT_FG="7F6000"
C_DEL="F4CCCC"; C_DEL_FG="990000"
C_BLUE="CFE2F3"; C_YELLOW="FFF9C4"; C_GREEN="E8F5E9"
C_ODD="FFFFFF"; C_EVEN="F5F5F5"
C_STEP="E8EAF6"; C_ANNO="FFF3E0"; C_CODE="F1F8E9"
MBG={"GET":C_GET,"POST":C_POST,"PUT":C_PUT,"DELETE":C_DEL}
MFG={"GET":C_GET_FG,"POST":C_POST_FG,"PUT":C_PUT_FG,"DELETE":C_DEL_FG}

def bd():
    s=Side(style="thin",color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)
def hf(sz=10,bold=True,color=C_WHITE):
    return Font(name="맑은 고딕",bold=bold,size=sz,color=color)
def bf(sz=9,bold=False,color="111111"):
    return Font(name="맑은 고딕",bold=bold,size=sz,color=color)
def cf(sz=8):
    return Font(name="Consolas",size=sz,color="1A237E")
def fl(c):
    return PatternFill("solid",fgColor=c)
def al(h="left",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def title(ws,row,text,end_col="J",height=30):
    ws.merge_cells(f"A{row}:{end_col}{row}")
    c=ws[f"A{row}"]
    c.value=text; c.font=hf(14); c.fill=fl(C_NAVY)
    c.alignment=al("center"); ws.row_dimensions[row].height=height
def hrow(ws,row,headers,height=22):
    for ci,h in enumerate(headers,1):
        c=ws.cell(row=row,column=ci,value=h)
        c.font=hf(9); c.fill=fl(C_DARK); c.alignment=al("center"); c.border=bd()
    ws.row_dimensions[row].height=height
def cw(ws,widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
def wc(ws,row,col,val,font=None,bg=C_ODD,halign="left",wrap=True,bold=False):
    c=ws.cell(row=row,column=col,value=val)
    c.font=font or bf(bold=bold); c.fill=fl(bg)
    c.alignment=al(halign,wrap=wrap); c.border=bd()
    return c
def mc(ws,row,col,method):
    c=ws.cell(row=row,column=col,value=method)
    c.font=Font(name="맑은 고딕",bold=True,size=9,color=MFG.get(method,"000000"))
    c.fill=fl(MBG.get(method,C_ODD)); c.alignment=al("center"); c.border=bd()


# ══════════════════════════════════════════════════════════
# 시트 1 - 코딩 순서 & 연결구조
# ══════════════════════════════════════════════════════════
ws0=wb.active; ws0.title="코딩 순서 & 연결구조"
title(ws0,1,"LOV3 프로젝트 - 코딩 순서 & 클래스 연결 구조","F")
cw(ws0,[6,28,40,30,20,20])

ws0.merge_cells("A2:F2")
c=ws0["A2"]; c.value="[ 클래스 연결 구조 ]  클라이언트 → Controller → Service 인터페이스 → ServiceImpl → Repository → DB"
c.font=Font(name="맑은 고딕",bold=True,size=11,color=C_NAVY)
c.fill=fl("E3F2FD"); c.alignment=al("center"); ws0.row_dimensions[2].height=20

flow=[
    ("","클라이언트 (브라우저/앱)","HTTP 요청 전송 (JSON 바디, 헤더에 JWT 토큰)","","",""),
    ("","    ↕  HTTP 요청/응답","","","",""),
    ("","Controller","@RestController  |  요청 받아 Service 호출 후 ResponseEntity로 응답","파일: *Controller.java","4단계","@RestController"),
    ("","    ↕  메서드 호출","","","",""),
    ("","Service 인터페이스","메서드 목록 정의  |  실제 구현은 ServiceImpl 담당","파일: *Service.java","2단계","interface"),
    ("","    ↕  Spring이 Impl 자동 연결","","","",""),
    ("","ServiceImpl","@Service  |  실제 핵심 로직 작성  |  Repository로 DB 접근","파일: *ServiceImpl.java","3단계","@Service @Transactional"),
    ("","    ↕  DB 접근","","","",""),
    ("","Repository","extends JpaRepository  |  이미 완성됨 (수정 불필요)","파일: *Repository.java","완성됨","extends JpaRepository"),
    ("","    ↕  SQL 자동 생성","","","",""),
    ("","MySQL DB","docker 컨테이너  |  ddl-auto=create 로 테이블 자동 생성","","",""),
]
for ri,rd in enumerate(flow,3):
    is_class=rd[1] in ("Controller","Service 인터페이스","ServiceImpl","Repository","MySQL DB","클라이언트 (브라우저/앱)")
    bg="E8F5E9" if is_class else C_ODD
    for ci,val in enumerate(rd,1):
        c=ws0.cell(row=ri,column=ci,value=val)
        c.font=bf(bold=is_class); c.fill=fl(bg)
        c.alignment=al(wrap=True); c.border=bd()
    ws0.row_dimensions[ri].height=28

ri=15
ws0.merge_cells(f"A{ri}:F{ri}")
c=ws0[f"A{ri}"]; c.value="[ 코딩 순서 ]  아래 순서대로 진행하세요"
c.font=Font(name="맑은 고딕",bold=True,size=11,color=C_NAVY)
c.fill=fl("FFF9C4"); c.alignment=al("center"); ws0.row_dimensions[ri].height=20; ri+=1
hrow(ws0,ri,["순서","파일명","해야할 일","참고 시트","단계 이름","핵심 키워드"]); ri+=1

steps=[
    ("1","MemberException.java\nBoardException.java",
     "① private final HttpStatus httpStatus; 필드 추가\n② 생성자 2개 추가 (message만 / message+HttpStatus)\n③ getHttpStatus() getter 추가\n④ import org.springframework.http.HttpStatus; 추가\n⑤ 완성 코드는 '에러 처리 가이드' 시트 참고",
     "에러 처리 가이드","Exception 완성","extends RuntimeException"),

    ("2","MemberService.java\nBoardService.java\nPlaceService.java\nDateRequestService.java",
     "① void → 올바른 return type으로 변경\n   예) void getMyInfo() → MemberResDto getMyInfo(Long memberNo)\n② () 안에 파라미터 추가\n③ 필요한 import 추가\n※ '클래스 작성법' 시트 참고",
     "클래스 작성법","인터페이스 시그니처","interface"),

    ("3-A","MemberServiceImpl.java",
     "① 클래스 선언 변경:\n   @Service @RequiredArgsConstructor @Transactional\n   public class MemberServiceImpl implements MemberService\n② 필드 추가:\n   private final MemberRepository memberRepository;\n   private final BCryptPasswordEncoder passwordEncoder;\n③ 각 메서드 @Override, return type, 파라미터 추가 후 로직 구현\n※ 'Service 구현 가이드' 시트 참고",
     "Service 구현 가이드","ServiceImpl 구현","@Service implements"),

    ("3-B","BoardServiceImpl.java",
     "① implements BoardService + @Service @RequiredArgsConstructor @Transactional\n② 필드: BoardRepository, MemberRepository, PlaceRepository",
     "Service 구현 가이드","ServiceImpl 구현","@Service"),

    ("3-C","PlaceServiceImpl.java",
     "① implements PlaceService + @Service @RequiredArgsConstructor @Transactional\n② 필드: PlaceRepository",
     "Service 구현 가이드","ServiceImpl 구현","@Service"),

    ("3-D","DateRequestServiceImpl.java",
     "① implements DateRequestService + @Service @RequiredArgsConstructor @Transactional\n② 필드: DateRequestRepository, BoardRepository, MemberRepository",
     "Service 구현 가이드","ServiceImpl 구현","@Service"),

    ("4-A","MemberController.java",
     "① @RestController @RequestMapping(\"/members\") @RequiredArgsConstructor\n② private final MemberService memberService;\n③ 각 메서드에 @GetMapping/@PostMapping 등 추가\n④ void → ResponseEntity<MemberResDto> 로 변경\n※ '클래스 작성법' 시트 참고",
     "클래스 작성법","Controller 구현","@RestController ResponseEntity"),

    ("4-B","AdminController.java","① @RestController @RequestMapping(\"/admin\") @RequiredArgsConstructor\n② private final MemberService memberService;","클래스 작성법","Controller 구현","@RestController"),
    ("4-C","BoardController.java","① @RestController @RequestMapping(\"/boards\") @RequiredArgsConstructor\n② private final BoardService boardService;","클래스 작성법","Controller 구현","@RestController"),
    ("4-D","PlaceController.java","① @RestController @RequestMapping(\"/places\") @RequiredArgsConstructor\n② private final PlaceService placeService;","클래스 작성법","Controller 구현","@RestController"),
    ("4-E","DateRequestController.java","① @RestController @RequiredArgsConstructor  (RequestMapping 없음)\n② private final DateRequestService dateRequestService;","클래스 작성법","Controller 구현","@RestController"),

    ("5","GlobalExceptionAdvice.java",
     "① @RestControllerAdvice 클래스에 추가\n② handleMemberException():\n   @ExceptionHandler(MemberException.class)\n   파라미터: (MemberException e)\n   return type: ResponseEntity<Map<String, String>>\n③ handleBoardException(): 동일 구조\n※ '에러 처리 가이드' 시트 참고",
     "에러 처리 가이드","예외 핸들러 완성","@RestControllerAdvice @ExceptionHandler"),
]

for sd in steps:
    step,files,todo,ref,name,anno=sd
    bg=C_STEP if step in("1","5") else(C_ANNO if step=="2" else C_CODE)
    wc(ws0,ri,1,step,bold=True,bg=bg,halign="center")
    wc(ws0,ri,2,files,bold=True,bg=bg)
    wc(ws0,ri,3,todo,bg=bg)
    wc(ws0,ri,4,ref,bg=bg,halign="center")
    wc(ws0,ri,5,name,bold=True,bg=bg,halign="center")
    c6=ws0.cell(row=ri,column=6,value=anno)
    c6.font=cf(8); c6.fill=fl(bg); c6.alignment=al(wrap=True); c6.border=bd()
    ws0.row_dimensions[ri].height=90; ri+=1

ws0.freeze_panes="A17"


# ══════════════════════════════════════════════════════════
# 시트 2 - 클래스 작성법
# ══════════════════════════════════════════════════════════
ws_cls=wb.create_sheet("클래스 작성법")
title(ws_cls,1,"클래스별 메서드 시그니처 & 어노테이션 작성법","F")
cw(ws_cls,[22,10,34,48,30,22])
hrow(ws_cls,2,["파일명 / 메서드명","HTTP","추가할 어노테이션","올바른 메서드 시그니처 (void 대신 이것으로)","파라미터 설명","return type"])

cls_rows=[
    # Exception
    ("[ MemberException.java ]","","extends RuntimeException 유지\nimport org.springframework.http.HttpStatus;",
     "private final HttpStatus httpStatus;\n\npublic MemberException(String message) {\n    super(message);\n    this.httpStatus = HttpStatus.BAD_REQUEST;\n}\npublic MemberException(String message, HttpStatus httpStatus) {\n    super(message);\n    this.httpStatus = httpStatus;\n}\npublic HttpStatus getHttpStatus() { return httpStatus; }","",""),
    ("[ BoardException.java ]","","MemberException과 동일한 구조",
     "BoardException으로 클래스명만 바꿔서 동일하게 작성","",""),

    # GlobalExceptionAdvice
    ("GlobalExceptionAdvice\n클래스 레벨","",
     "@RestControllerAdvice\nimport ResponseEntity, ExceptionHandler\nimport MemberException, BoardException\nimport HashMap, Map",
     "public class GlobalExceptionAdvice { ... }","",""),
    ("handleMemberException()","","@ExceptionHandler(MemberException.class)",
     "public ResponseEntity<Map<String, String>> handleMemberException(MemberException e)",
     "MemberException e","ResponseEntity<Map<String,String>>"),
    ("handleBoardException()","","@ExceptionHandler(BoardException.class)",
     "public ResponseEntity<Map<String, String>> handleBoardException(BoardException e)",
     "BoardException e","ResponseEntity<Map<String,String>>"),

    # MemberService
    ("[ MemberService.java ]\nsignUp()","POST","interface 메서드",
     "MemberResDto signUp(MemberReqDto dto);","MemberReqDto dto","MemberResDto"),
    ("checkDuplicateId()","GET","interface 메서드",
     "boolean checkDuplicateId(String id);","String id","boolean"),
    ("getMyInfo()","GET","interface 메서드",
     "MemberResDto getMyInfo(Long memberNo);","Long memberNo","MemberResDto"),
    ("updateMyInfo()","PUT","interface 메서드",
     "MemberResDto updateMyInfo(Long memberNo, MemberReqDto dto);","MemberReqDto에 info+profile 필드 모두 포함","MemberResDto"),
    ("deleteMyAccount()","DELETE","interface 메서드",
     "void deleteMyAccount(Long memberNo);","Long memberNo","void"),
    ("getAllMembers()","GET","interface 메서드",
     "List<MemberResDto> getAllMembers();","","List<MemberResDto>"),
    ("deleteMemberByAdmin()","DELETE","interface 메서드",
     "void deleteMemberByAdmin(Long memberNo);","Long memberNo","void"),

    # BoardService
    ("[ BoardService.java ]\ngetAllBoards()","GET","interface 메서드","List<BoardResDto> getAllBoards();","","List<BoardResDto>"),
    ("getBoard()","GET","interface 메서드","BoardResDto getBoard(Long boardId);","Long boardId","BoardResDto"),
    ("searchBoards()","GET","interface 메서드","List<BoardResDto> searchBoards(String keyword);","String keyword","List<BoardResDto>"),
    ("getBoardsByPlace()","GET","interface 메서드","List<BoardResDto> getBoardsByPlace(Long placeId);","Long placeId","List<BoardResDto>"),
    ("addBoard()","POST","interface 메서드","BoardResDto addBoard(BoardReqDto dto, Long memberNo);","","BoardResDto"),
    ("updateBoard()","PUT","interface 메서드","BoardResDto updateBoard(Long boardId, BoardReqDto dto, Long memberNo);","","BoardResDto"),
    ("deleteBoard()","DELETE","interface 메서드","void deleteBoard(Long boardId, Long memberNo, String role);","","void"),

    # PlaceService
    ("[ PlaceService.java ]\ngetAllPlaces()","GET","interface 메서드","List<PlaceResDto> getAllPlaces();","","List<PlaceResDto>"),
    ("searchPlaces()","GET","interface 메서드","List<PlaceResDto> searchPlaces(String keyword);","","List<PlaceResDto>"),
    ("getPlace()","GET","interface 메서드","PlaceResDto getPlace(Long placeId);","","PlaceResDto"),
    ("addPlace()","POST","interface 메서드","PlaceResDto addPlace(PlaceReqDto dto);","","PlaceResDto"),
    ("updatePlace()","PUT","interface 메서드","PlaceResDto updatePlace(Long placeId, PlaceReqDto dto);","","PlaceResDto"),
    ("deletePlace()","DELETE","interface 메서드","void deletePlace(Long placeId);","","void"),

    # DateRequestService
    ("[ DateRequestService.java ]\nsendRequest()","POST","interface 메서드",
     "DateRequestResDto sendRequest(Long boardId, DateRequestReqDto dto, Long senderNo);","","DateRequestResDto"),
    ("getReceivedRequestsByBoard()","GET","interface 메서드",
     "List<DateRequestResDto> getReceivedRequestsByBoard(Long boardId, Long memberNo);","","List<DateRequestResDto>"),
    ("getAllReceivedRequests()","GET","interface 메서드",
     "List<DateRequestResDto> getAllReceivedRequests(Long memberNo);","","List<DateRequestResDto>"),
    ("getSentRequests()","GET","interface 메서드",
     "List<DateRequestResDto> getSentRequests(Long memberNo);","","List<DateRequestResDto>"),
    ("acceptRequest()","PUT","interface 메서드",
     "DateRequestResDto acceptRequest(Long requestId, Long memberNo);","","DateRequestResDto"),
    ("rejectRequest()","PUT","interface 메서드",
     "DateRequestResDto rejectRequest(Long requestId, Long memberNo);","","DateRequestResDto"),
    ("cancelRequest()","DELETE","interface 메서드",
     "void cancelRequest(Long requestId, Long memberNo);","","void"),

    # MemberServiceImpl
    ("[ MemberServiceImpl.java ]\n클래스 레벨","",
     "@Service @RequiredArgsConstructor @Transactional\npublic class MemberServiceImpl implements MemberService",
     "private final MemberRepository memberRepository;\nprivate final BCryptPasswordEncoder passwordEncoder;",
     "import: Service, RequiredArgsConstructor, Transactional\nBCryptPasswordEncoder, Member\nMemberReqDto, MemberResDto\nMemberException, MemberRepository",""),
    ("signUp() @Override","POST","@Override","public MemberResDto signUp(MemberReqDto dto)","","MemberResDto"),
    ("checkDuplicateId() @Override","GET","@Override @Transactional(readOnly=true)","public boolean checkDuplicateId(String id)","","boolean"),
    ("getMyInfo() @Override","GET","@Override @Transactional(readOnly=true)","public MemberResDto getMyInfo(Long memberNo)","","MemberResDto"),
    ("updateMyInfo() @Override","PUT","@Override","public MemberResDto updateMyInfo(Long memberNo, MemberReqDto dto)","MemberReqDto에 info+profile 필드 모두 포함","MemberResDto"),
    ("deleteMyAccount() @Override","DELETE","@Override","public void deleteMyAccount(Long memberNo)","","void"),
    ("getAllMembers() @Override","GET","@Override @Transactional(readOnly=true)","public List<MemberResDto> getAllMembers()","","List<MemberResDto>"),
    ("deleteMemberByAdmin() @Override","DELETE","@Override","public void deleteMemberByAdmin(Long memberNo)","","void"),

    # MemberController
    ("[ MemberController.java ]\n클래스 레벨","",
     "@RestController @RequestMapping(\"/members\") @RequiredArgsConstructor\npublic class MemberController",
     "private final MemberService memberService;",
     "import: RestController, RequestMapping, RequiredArgsConstructor\nHttpStatus, ResponseEntity, AuthenticationPrincipal\nRequestBody, RequestParam, PathVariable\nMemberReqDto, MemberResDto\nCustomMemberDetails, MemberService",""),
    ("signUp()","POST","@PostMapping",
     "public ResponseEntity<MemberResDto> signUp(@RequestBody MemberReqDto dto)",
     "@RequestBody MemberReqDto dto","ResponseEntity<MemberResDto>"),
    ("checkDuplicateId()","GET","@GetMapping(\"/check\")",
     "public ResponseEntity<Boolean> checkDuplicateId(@RequestParam String id)",
     "@RequestParam String id","ResponseEntity<Boolean>"),
    ("getMyInfo()","GET","@GetMapping(\"/me\")",
     "public ResponseEntity<MemberResDto> getMyInfo(@AuthenticationPrincipal CustomMemberDetails memberDetails)",
     "@AuthenticationPrincipal CustomMemberDetails memberDetails","ResponseEntity<MemberResDto>"),
    ("updateMyInfo()","PUT","@PutMapping(\"/me\")",
     "public ResponseEntity<MemberResDto> updateMyInfo(@AuthenticationPrincipal CustomMemberDetails memberDetails, @RequestBody MemberReqDto dto)",
     "info+profile 필드 모두 MemberReqDto 하나로 받음","ResponseEntity<MemberResDto>"),
    ("deleteMyAccount()","DELETE","@DeleteMapping(\"/me\")",
     "public ResponseEntity<Void> deleteMyAccount(@AuthenticationPrincipal CustomMemberDetails memberDetails)",
     "","ResponseEntity<Void>"),

    # AdminController
    ("[ AdminController.java ]\n클래스 레벨","",
     "@RestController @RequestMapping(\"/admin\") @RequiredArgsConstructor\npublic class AdminController",
     "private final MemberService memberService;","",""),
    ("getAllMembers()","GET","@GetMapping(\"/members\")",
     "public ResponseEntity<List<MemberResDto>> getAllMembers()","없음","ResponseEntity<List<MemberResDto>>"),
    ("deleteMember()","DELETE","@DeleteMapping(\"/members/{memberNo}\")",
     "public ResponseEntity<Void> deleteMember(@PathVariable Long memberNo)",
     "@PathVariable Long memberNo","ResponseEntity<Void>"),

    # BoardServiceImpl
    ("[ BoardServiceImpl.java ]\n클래스 레벨","",
     "@Service @RequiredArgsConstructor @Transactional\npublic class BoardServiceImpl implements BoardService",
     "private final BoardRepository boardRepository;\nprivate final MemberRepository memberRepository;\nprivate final PlaceRepository placeRepository;","",""),
    ("getAllBoards() @Override","GET","@Override @Transactional(readOnly=true)","public List<BoardResDto> getAllBoards()","","List<BoardResDto>"),
    ("getBoard() @Override","GET","@Override","public BoardResDto getBoard(Long boardId)","","BoardResDto"),
    ("searchBoards() @Override","GET","@Override @Transactional(readOnly=true)","public List<BoardResDto> searchBoards(String keyword)","","List<BoardResDto>"),
    ("getBoardsByPlace() @Override","GET","@Override @Transactional(readOnly=true)","public List<BoardResDto> getBoardsByPlace(Long placeId)","","List<BoardResDto>"),
    ("addBoard() @Override","POST","@Override","public BoardResDto addBoard(BoardReqDto dto, Long memberNo)","","BoardResDto"),
    ("updateBoard() @Override","PUT","@Override","public BoardResDto updateBoard(Long boardId, BoardReqDto dto, Long memberNo)","","BoardResDto"),
    ("deleteBoard() @Override","DELETE","@Override","public void deleteBoard(Long boardId, Long memberNo, String role)","","void"),

    # BoardController
    ("[ BoardController.java ]\n클래스 레벨","",
     "@RestController @RequestMapping(\"/boards\") @RequiredArgsConstructor\npublic class BoardController",
     "private final BoardService boardService;","",""),
    ("getAllBoards()","GET","@GetMapping","public ResponseEntity<List<BoardResDto>> getAllBoards()","","ResponseEntity<List<BoardResDto>>"),
    ("getBoard()","GET","@GetMapping(\"/{boardId}\")","public ResponseEntity<BoardResDto> getBoard(@PathVariable Long boardId)","","ResponseEntity<BoardResDto>"),
    ("searchBoards()","GET","@GetMapping(\"/search\")","public ResponseEntity<List<BoardResDto>> searchBoards(@RequestParam String keyword)","","ResponseEntity<List<BoardResDto>>"),
    ("getBoardsByPlace()","GET","@GetMapping(\"/place/{placeId}\")","public ResponseEntity<List<BoardResDto>> getBoardsByPlace(@PathVariable Long placeId)","","ResponseEntity<List<BoardResDto>>"),
    ("addBoard()","POST","@PostMapping","public ResponseEntity<BoardResDto> addBoard(@AuthenticationPrincipal CustomMemberDetails memberDetails, @RequestBody BoardReqDto dto)","","ResponseEntity<BoardResDto>"),
    ("updateBoard()","PUT","@PutMapping(\"/{boardId}\")","public ResponseEntity<BoardResDto> updateBoard(@PathVariable Long boardId, @AuthenticationPrincipal CustomMemberDetails memberDetails, @RequestBody BoardReqDto dto)","","ResponseEntity<BoardResDto>"),
    ("deleteBoard()","DELETE","@DeleteMapping(\"/{boardId}\")","public ResponseEntity<Void> deleteBoard(@PathVariable Long boardId, @AuthenticationPrincipal CustomMemberDetails memberDetails)","","ResponseEntity<Void>"),

    # PlaceServiceImpl
    ("[ PlaceServiceImpl.java ]\n클래스 레벨","",
     "@Service @RequiredArgsConstructor @Transactional\npublic class PlaceServiceImpl implements PlaceService",
     "private final PlaceRepository placeRepository;","",""),
    ("getAllPlaces() @Override","GET","@Override @Transactional(readOnly=true)","public List<PlaceResDto> getAllPlaces()","","List<PlaceResDto>"),
    ("searchPlaces() @Override","GET","@Override @Transactional(readOnly=true)","public List<PlaceResDto> searchPlaces(String keyword)","","List<PlaceResDto>"),
    ("getPlace() @Override","GET","@Override @Transactional(readOnly=true)","public PlaceResDto getPlace(Long placeId)","","PlaceResDto"),
    ("addPlace() @Override","POST","@Override","public PlaceResDto addPlace(PlaceReqDto dto)","","PlaceResDto"),
    ("updatePlace() @Override","PUT","@Override","public PlaceResDto updatePlace(Long placeId, PlaceReqDto dto)","","PlaceResDto"),
    ("deletePlace() @Override","DELETE","@Override","public void deletePlace(Long placeId)","","void"),

    # PlaceController
    ("[ PlaceController.java ]\n클래스 레벨","",
     "@RestController @RequestMapping(\"/places\") @RequiredArgsConstructor\npublic class PlaceController",
     "private final PlaceService placeService;","",""),
    ("getAllPlaces()","GET","@GetMapping","public ResponseEntity<List<PlaceResDto>> getAllPlaces()","","ResponseEntity<List<PlaceResDto>>"),
    ("searchPlaces()","GET","@GetMapping(\"/search\")","public ResponseEntity<List<PlaceResDto>> searchPlaces(@RequestParam String keyword)","","ResponseEntity<List<PlaceResDto>>"),
    ("getPlace()","GET","@GetMapping(\"/{placeId}\")","public ResponseEntity<PlaceResDto> getPlace(@PathVariable Long placeId)","","ResponseEntity<PlaceResDto>"),
    ("addPlace()","POST","@PostMapping","public ResponseEntity<PlaceResDto> addPlace(@RequestBody PlaceReqDto dto)","","ResponseEntity<PlaceResDto>"),
    ("updatePlace()","PUT","@PutMapping(\"/{placeId}\")","public ResponseEntity<PlaceResDto> updatePlace(@PathVariable Long placeId, @RequestBody PlaceReqDto dto)","","ResponseEntity<PlaceResDto>"),
    ("deletePlace()","DELETE","@DeleteMapping(\"/{placeId}\")","public ResponseEntity<Void> deletePlace(@PathVariable Long placeId)","","ResponseEntity<Void>"),

    # DateRequestServiceImpl
    ("[ DateRequestServiceImpl.java ]\n클래스 레벨","",
     "@Service @RequiredArgsConstructor @Transactional\npublic class DateRequestServiceImpl implements DateRequestService",
     "private final DateRequestRepository dateRequestRepository;\nprivate final BoardRepository boardRepository;\nprivate final MemberRepository memberRepository;","",""),
    ("sendRequest() @Override","POST","@Override","public DateRequestResDto sendRequest(Long boardId, DateRequestReqDto dto, Long senderNo)","","DateRequestResDto"),
    ("getReceivedRequestsByBoard() @Override","GET","@Override @Transactional(readOnly=true)","public List<DateRequestResDto> getReceivedRequestsByBoard(Long boardId, Long memberNo)","","List<DateRequestResDto>"),
    ("getAllReceivedRequests() @Override","GET","@Override @Transactional(readOnly=true)","public List<DateRequestResDto> getAllReceivedRequests(Long memberNo)","","List<DateRequestResDto>"),
    ("getSentRequests() @Override","GET","@Override @Transactional(readOnly=true)","public List<DateRequestResDto> getSentRequests(Long memberNo)","","List<DateRequestResDto>"),
    ("acceptRequest() @Override","PUT","@Override","public DateRequestResDto acceptRequest(Long requestId, Long memberNo)","","DateRequestResDto"),
    ("rejectRequest() @Override","PUT","@Override","public DateRequestResDto rejectRequest(Long requestId, Long memberNo)","","DateRequestResDto"),
    ("cancelRequest() @Override","DELETE","@Override","public void cancelRequest(Long requestId, Long memberNo)","","void"),

    # DateRequestController
    ("[ DateRequestController.java ]\n클래스 레벨","",
     "@RestController @RequiredArgsConstructor\npublic class DateRequestController\n※ @RequestMapping 없음",
     "private final DateRequestService dateRequestService;","",""),
    ("sendRequest()","POST","@PostMapping(\"/boards/{boardId}/date-requests\")","public ResponseEntity<DateRequestResDto> sendRequest(@PathVariable Long boardId, @AuthenticationPrincipal CustomMemberDetails memberDetails, @RequestBody DateRequestReqDto dto)","","ResponseEntity<DateRequestResDto>"),
    ("getReceivedRequestsByBoard()","GET","@GetMapping(\"/boards/{boardId}/date-requests\")","public ResponseEntity<List<DateRequestResDto>> getReceivedRequestsByBoard(@PathVariable Long boardId, @AuthenticationPrincipal CustomMemberDetails memberDetails)","","ResponseEntity<List<DateRequestResDto>>"),
    ("getAllReceivedRequests()","GET","@GetMapping(\"/date-requests/received\")","public ResponseEntity<List<DateRequestResDto>> getAllReceivedRequests(@AuthenticationPrincipal CustomMemberDetails memberDetails)","","ResponseEntity<List<DateRequestResDto>>"),
    ("getSentRequests()","GET","@GetMapping(\"/date-requests/sent\")","public ResponseEntity<List<DateRequestResDto>> getSentRequests(@AuthenticationPrincipal CustomMemberDetails memberDetails)","","ResponseEntity<List<DateRequestResDto>>"),
    ("acceptRequest()","PUT","@PutMapping(\"/date-requests/{requestId}/accept\")","public ResponseEntity<DateRequestResDto> acceptRequest(@PathVariable Long requestId, @AuthenticationPrincipal CustomMemberDetails memberDetails)","","ResponseEntity<DateRequestResDto>"),
    ("rejectRequest()","PUT","@PutMapping(\"/date-requests/{requestId}/reject\")","public ResponseEntity<DateRequestResDto> rejectRequest(@PathVariable Long requestId, @AuthenticationPrincipal CustomMemberDetails memberDetails)","","ResponseEntity<DateRequestResDto>"),
    ("cancelRequest()","DELETE","@DeleteMapping(\"/date-requests/{requestId}\")","public ResponseEntity<Void> cancelRequest(@PathVariable Long requestId, @AuthenticationPrincipal CustomMemberDetails memberDetails)","","ResponseEntity<Void>"),
]

cls_colors={
    "[ MemberException":"FCE4EC","[ BoardException":"FCE4EC",
    "GlobalException":"F3E5F5",
    "[ MemberService":"E3F2FD","[ BoardService":"E3F2FD",
    "[ PlaceService":"E3F2FD","[ DateRequestService":"E3F2FD",
    "[ MemberServiceImpl":"E8F5E9","[ BoardServiceImpl":"E8F5E9",
    "[ PlaceServiceImpl":"E8F5E9","[ DateRequestServiceImpl":"E8F5E9",
    "[ MemberController":"FFF8E1","[ AdminController":"FFF8E1",
    "[ BoardController":"FFF8E1","[ PlaceController":"FFF8E1",
    "[ DateRequestController":"FFF8E1",
}

for ri_idx,row in enumerate(cls_rows,3):
    name,method,anno,sig,param,ret=row
    bg=C_ODD
    for k,v in cls_colors.items():
        if name.startswith(k): bg=v; break
    is_header=name.startswith("[")
    c1=ws_cls.cell(row=ri_idx,column=1,value=name)
    if is_header: c1.font=hf(10); c1.fill=fl(C_NAVY)
    else: c1.font=bf(sz=9); c1.fill=fl(bg)
    c1.alignment=al(wrap=True); c1.border=bd()
    if method: mc(ws_cls,ri_idx,2,method)
    else: wc(ws_cls,ri_idx,2,"",bg=bg if not is_header else C_NAVY)
    c3=ws_cls.cell(row=ri_idx,column=3,value=anno)
    c3.font=cf(8); c3.fill=fl(bg); c3.alignment=al(wrap=True); c3.border=bd()
    c4=ws_cls.cell(row=ri_idx,column=4,value=sig)
    c4.font=cf(8); c4.fill=fl(bg); c4.alignment=al(wrap=True); c4.border=bd()
    wc(ws_cls,ri_idx,5,param,bg=bg)
    wc(ws_cls,ri_idx,6,ret,bg=bg,halign="center")
    ws_cls.row_dimensions[ri_idx].height=60 if is_header else 36

ws_cls.freeze_panes="A3"


# ══════════════════════════════════════════════════════════
# 시트 3 - API 전체목록 (29개)
# ══════════════════════════════════════════════════════════
ws=wb.create_sheet("API 전체목록")
title(ws,1,"LOV3 장소 기반 데이트 매칭 게시판 - REST API 전체 목록 (28개)","K")
ws.merge_cells("A2:K2")
c=ws["A2"]
c.value="서버: http://localhost:8087  |  인증: JWT Bearer Token  |  Content-Type: application/json"
c.font=bf(sz=10,color="555555"); c.alignment=al("center"); ws.row_dimensions[2].height=16
heads=["No","도메인","메서드","URL","기능 설명","인증","Request Body","성공코드","컨트롤러 메서드","서비스 메서드","비고"]
cw(ws,[4,10,8,44,26,6,12,13,26,30,22])
hrow(ws,3,heads)

apis=[
    (1,"인증","POST","/login","로그인 (JWT 발급)","X","O","200 OK","LoginFilter (자동처리)","-","응답 헤더에 토큰"),
    (2,"회원","POST","/members","회원가입","X","O","201 Created","signUp(dto)","signUp(dto)",""),
    (3,"회원","GET","/members/check?id={id}","ID 중복 확인","X","X","200 OK","checkDuplicateId(id)","checkDuplicateId(id)","true=사용가능"),
    (4,"회원","GET","/members/me","내 정보 조회","O","X","200 OK","getMyInfo(memberDetails)","getMyInfo(memberNo)","MemberResDto 반환"),
    (5,"회원","PUT","/members/me","내 정보 + 프로필 수정","O","O","200 OK","updateMyInfo(...)","updateMyInfo(memberNo,dto)","이름/이메일/비번/사진/성별/나이/소개"),
    (6,"회원","DELETE","/members/me","회원 탈퇴","O","X","204 No Content","deleteMyAccount(...)","deleteMyAccount(memberNo)",""),
    (7,"관리자","GET","/admin/members","전체 회원 목록","O","X","200 OK","getAllMembers()","getAllMembers()","ROLE_ADMIN"),
    (8,"관리자","DELETE","/admin/members/{memberNo}","회원 강제 탈퇴","O","X","204 No Content","deleteMember(memberNo)","deleteMemberByAdmin(memberNo)","ROLE_ADMIN"),
    (9,"장소","GET","/places","장소 전체 목록","X","X","200 OK","getAllPlaces()","getAllPlaces()",""),
    (10,"장소","GET","/places/search?keyword={keyword}","장소 검색","X","X","200 OK","searchPlaces(keyword)","searchPlaces(keyword)",""),
    (11,"장소","GET","/places/{placeId}","장소 단건 조회","X","X","200 OK","getPlace(placeId)","getPlace(placeId)",""),
    (12,"장소","POST","/places","장소 등록","O","O","201 Created","addPlace(dto)","addPlace(dto)","ROLE_ADMIN"),
    (13,"장소","PUT","/places/{placeId}","장소 수정","O","O","200 OK","updatePlace(placeId,dto)","updatePlace(placeId,dto)","ROLE_ADMIN"),
    (14,"장소","DELETE","/places/{placeId}","장소 삭제","O","X","204 No Content","deletePlace(placeId)","deletePlace(placeId)","ROLE_ADMIN"),
    (15,"게시글","GET","/boards","게시글 전체 목록","X","X","200 OK","getAllBoards()","getAllBoards()","장소 정보 포함"),
    (16,"게시글","GET","/boards/{boardId}","게시글 단건 조회","X","X","200 OK","getBoard(boardId)","getBoard(boardId)","조회수 +1"),
    (17,"게시글","GET","/boards/search?keyword={keyword}","게시글 검색","X","X","200 OK","searchBoards(keyword)","searchBoards(keyword)",""),
    (18,"게시글","GET","/boards/place/{placeId}","특정 장소 게시글 목록","X","X","200 OK","getBoardsByPlace(placeId)","getBoardsByPlace(placeId)",""),
    (19,"게시글","POST","/boards","게시글 등록","O","O","201 Created","addBoard(...)","addBoard(dto,memberNo)","placeId 필수"),
    (20,"게시글","PUT","/boards/{boardId}","게시글 수정","O","O","200 OK","updateBoard(...)","updateBoard(boardId,dto,...)","본인만"),
    (21,"게시글","DELETE","/boards/{boardId}","게시글 삭제","O","X","204 No Content","deleteBoard(...)","deleteBoard(boardId,...)","본인/관리자"),
    (22,"데이트신청","POST","/boards/{boardId}/date-requests","데이트 신청 보내기","O","O","201 Created","sendRequest(...)","sendRequest(boardId,dto,...)","본인 글 신청 불가"),
    (23,"데이트신청","GET","/boards/{boardId}/date-requests","특정 게시글 신청 목록","O","X","200 OK","getReceivedRequestsByBoard","getReceivedRequestsByBoard","게시글 작성자만"),
    (24,"데이트신청","GET","/date-requests/received","내가 받은 신청 전체","O","X","200 OK","getAllReceivedRequests(...)","getAllReceivedRequests(...)",""),
    (25,"데이트신청","GET","/date-requests/sent","내가 보낸 신청 목록","O","X","200 OK","getSentRequests(...)","getSentRequests(memberNo)",""),
    (26,"데이트신청","PUT","/date-requests/{requestId}/accept","신청 수락","O","X","200 OK","acceptRequest(...)","acceptRequest(requestId,...)","status→ACCEPTED"),
    (27,"데이트신청","PUT","/date-requests/{requestId}/reject","신청 거절","O","X","200 OK","rejectRequest(...)","rejectRequest(requestId,...)","status→REJECTED"),
    (28,"데이트신청","DELETE","/date-requests/{requestId}","신청 취소","O","X","204 No Content","cancelRequest(...)","cancelRequest(requestId,...)","신청 보낸 사람만"),
]

domain_colors={"인증":"E8EAF6","회원":"E3F2FD","관리자":"FCE4EC","장소":"E8F5E9","게시글":"FFF8E1","데이트신청":"F3E5F5"}
for ri,row in enumerate(apis,4):
    no,domain,method,url,desc,auth,body,code,ctrl,svc,note=row
    bg=domain_colors.get(domain,C_ODD)
    wc(ws,ri,1,no,bg=bg,halign="center",bold=True)
    wc(ws,ri,2,domain,bg=bg,halign="center")
    mc(ws,ri,3,method)
    wc(ws,ri,4,url,bg=bg)
    wc(ws,ri,5,desc,bg=bg)
    c=ws.cell(row=ri,column=6,value=auth)
    c.font=Font(name="맑은 고딕",bold=True,size=9,color="CC0000" if auth=="O" else "38761D")
    c.fill=fl(bg); c.alignment=al("center"); c.border=bd()
    wc(ws,ri,7,body,bg=bg,halign="center")
    wc(ws,ri,8,code,bg=bg,halign="center")
    wc(ws,ri,9,ctrl,bg=bg)
    wc(ws,ri,10,svc,bg=bg)
    wc(ws,ri,11,note,bg=bg)
    ws.row_dimensions[ri].height=17

ws.freeze_panes="A4"


# ══════════════════════════════════════════════════════════
# 시트 4 - DTO 구조
# ══════════════════════════════════════════════════════════
ws_dto=wb.create_sheet("DTO 구조")
title(ws_dto,1,"DTO (Data Transfer Object) 구조 정리","E")
cw(ws_dto,[20,14,42,36,18])
hrow(ws_dto,2,["DTO 파일명","방향","필드 목록","설명","어디서 사용"])

dtos=[
    ("MemberReqDto","요청 (→서버)",
     "String id\nString pwd\nString name\nString email\nString profileImg\nString gender\nInteger age\nString introduce",
     "회원가입 / 정보수정 / 프로필수정 시 클라이언트가 보내는 데이터\ngender: MALE / FEMALE / OTHER\n※ MemberProfileReqDto 제거 후 통합됨",
     "POST /members\nPUT /members/me"),
    ("MemberResDto","응답 (←서버)",
     "Long memberNo\nString id\nString name\nString email\nString profileImg\nString gender\nInteger age\nString introduce\nLocalDateTime regDate",
     "회원 정보 응답 DTO\n모든 회원 관련 응답에 이것 하나만 사용\n(예전 MemberProfileResDto 제거됨)",
     "모든 회원 응답"),
    ("BoardReqDto","요청 (→서버)",
     "String title\nString content\nLong placeId",
     "게시글 등록/수정 시 보내는 데이터\nplaceId 필수!",
     "POST /boards\nPUT /boards/{id}"),
    ("BoardResDto","응답 (←서버)",
     "Long boardId\nString title\nString content\nint viewCount\nMemberResDto member\nPlaceResDto place\nLocalDateTime regDate\nLocalDateTime updateDate",
     "게시글 응답\n작성자(MemberResDto)와 장소(PlaceResDto) 포함",
     "모든 게시글 응답"),
    ("PlaceReqDto","요청 (→서버)",
     "String placeName\nString placeImg\nString placeInfo",
     "장소 등록/수정 시 보내는 데이터 (관리자 전용)",
     "POST /places\nPUT /places/{id}"),
    ("PlaceResDto","응답 (←서버)",
     "Long placeId\nString placeName\nString placeImg\nString placeInfo",
     "장소 응답 DTO",
     "모든 장소 응답"),
    ("DateRequestReqDto","요청 (→서버)",
     "String message",
     "데이트 신청 시 보내는 데이터\n신청 멘트 하나만",
     "POST /boards/{id}/date-requests"),
    ("DateRequestResDto","응답 (←서버)",
     "Long requestId\nString message\nString status\nMemberResDto sender\nLong boardId\nLocalDateTime regDate",
     "데이트 신청 응답\nstatus: PENDING / ACCEPTED / REJECTED\nsender: 신청 보낸 사람 기본 정보",
     "모든 데이트신청 응답"),
]

for ri,row in enumerate(dtos,3):
    name,direction,fields,desc,usage=row
    bg=C_ODD if ri%2==1 else C_EVEN
    is_req="요청" in direction
    wc(ws_dto,ri,1,name,bold=True,bg=bg)
    c2=ws_dto.cell(row=ri,column=2,value=direction)
    c2.font=Font(name="맑은 고딕",bold=True,size=9,color="B45F06" if is_req else "38761D")
    c2.fill=fl(bg); c2.alignment=al("center",wrap=True); c2.border=bd()
    c3=ws_dto.cell(row=ri,column=3,value=fields)
    c3.font=cf(8); c3.fill=fl(bg); c3.alignment=al(wrap=True); c3.border=bd()
    wc(ws_dto,ri,4,desc,bg=bg)
    wc(ws_dto,ri,5,usage,bg=bg,halign="center")
    ws_dto.row_dimensions[ri].height=90

ws_dto.freeze_panes="A3"


# ══════════════════════════════════════════════════════════
# 시트 5 - Service 구현 가이드
# ══════════════════════════════════════════════════════════
ws4=wb.create_sheet("Service 구현 가이드")
title(ws4,1,"Service 구현 가이드 - ServiceImpl 메서드 안에 작성할 로직","C")
cw(ws4,[28,64,30])
hrow(ws4,2,["메서드명 (파일명)","메서드 안에 작성할 코드","핵심 설명"])

svc_items=[
    ("signUp\n(MemberServiceImpl)",
     'if (memberRepository.existsById(dto.getId())) {\n    throw new MemberException("이미 사용 중인 ID입니다.");\n}\nMember member = dto.toMember();\nmember.setPwd(passwordEncoder.encode(dto.getPwd()));\nmember.setRole("ROLE_USER");\nMember saved = memberRepository.save(member);\nreturn new MemberResDto(saved);',
     "비밀번호 암호화 필수!\nexistsById = 중복 확인"),
    ("checkDuplicateId\n(MemberServiceImpl)",
     'boolean exists = memberRepository.existsById(id);\nreturn !exists;',
     "true=사용가능\nfalse=중복"),
    ("getMyInfo\n(MemberServiceImpl)",
     'Member member = memberRepository.findById(memberNo)\n    .orElseThrow(() -> new MemberException("회원을 찾을 수 없습니다."));\nreturn new MemberResDto(member);',
     "orElseThrow = 없으면 예외"),
    ("updateMyInfo\n(MemberServiceImpl)",
     'Member member = memberRepository.findById(memberNo)\n    .orElseThrow(() -> new MemberException("없음"));\nmember.setName(dto.getName());\nmember.setEmail(dto.getEmail());\nif (dto.getPwd() != null && !dto.getPwd().isEmpty()) {\n    member.setPwd(passwordEncoder.encode(dto.getPwd()));\n}\nmember.setProfileImg(dto.getProfileImg());\nmember.setGender(dto.getGender());\nmember.setAge(dto.getAge());\nmember.setIntroduce(dto.getIntroduce());\nreturn new MemberResDto(member);',
     "info + 프로필 한번에 수정\n@Transactional → save() 불필요"),
    ("deleteMyAccount\n(MemberServiceImpl)",
     'Member member = memberRepository.findById(memberNo)\n    .orElseThrow(() -> new MemberException("없음"));\nmemberRepository.delete(member);',
     "void → return 없음"),
    ("getAllMembers\n(MemberServiceImpl)",
     'return memberRepository.findAll().stream()\n    .map(m -> new MemberResDto(m)).toList();',
     "관리자용"),
    ("deleteMemberByAdmin\n(MemberServiceImpl)",
     'Member member = memberRepository.findById(memberNo)\n    .orElseThrow(() -> new MemberException("없음"));\nmemberRepository.delete(member);',
     "관리자 강제 탈퇴"),
    ("getAllBoards\n(BoardServiceImpl)",
     'return boardRepository.findAllWithMemberAndPlace().stream()\n    .map(b -> new BoardResDto(b)).toList();',
     "fetch join으로 N+1 방지"),
    ("getBoard\n(BoardServiceImpl)",
     'Board board = boardRepository.findById(boardId)\n    .orElseThrow(() -> new BoardException("없음"));\nboard.setViewCount(board.getViewCount() + 1);\nreturn new BoardResDto(board);',
     "조회수 +1"),
    ("addBoard\n(BoardServiceImpl)",
     'Member member = memberRepository.findById(memberNo)\n    .orElseThrow(() -> new MemberException("없음"));\nPlace place = placeRepository.findById(dto.getPlaceId())\n    .orElseThrow(() -> new BoardException("장소 없음"));\nBoard board = new Board();\nboard.setTitle(dto.getTitle());\nboard.setContent(dto.getContent());\nboard.setMember(member);\nboard.setPlace(place);\nBoard saved = boardRepository.save(board);\nreturn new BoardResDto(saved);',
     "장소 set 필수!"),
    ("updateBoard\n(BoardServiceImpl)",
     'Board board = boardRepository.findById(boardId)\n    .orElseThrow(() -> new BoardException("없음"));\nif (!board.getMember().getMemberNo().equals(memberNo)) {\n    throw new BoardException("본인 게시글만 수정 가능");\n}\nboard.setTitle(dto.getTitle());\nboard.setContent(dto.getContent());\nif (dto.getPlaceId() != null) {\n    Place place = placeRepository.findById(dto.getPlaceId())\n        .orElseThrow(() -> new BoardException("장소 없음"));\n    board.setPlace(place);\n}\nreturn new BoardResDto(board);',
     "Long 비교는 .equals()"),
    ("deleteBoard\n(BoardServiceImpl)",
     'Board board = boardRepository.findById(boardId)\n    .orElseThrow(() -> new BoardException("없음"));\nboolean isOwner = board.getMember().getMemberNo().equals(memberNo);\nboolean isAdmin = "ROLE_ADMIN".equals(role);\nif (!isOwner && !isAdmin) throw new BoardException("권한 없음");\nboardRepository.delete(board);',
     "본인 OR 관리자"),
    ("getAllPlaces\n(PlaceServiceImpl)",
     'return placeRepository.findAll().stream()\n    .map(p -> new PlaceResDto(p)).toList();',""),
    ("addPlace\n(PlaceServiceImpl)",
     'Place place = new Place();\nplace.setPlaceName(dto.getPlaceName());\nplace.setPlaceImg(dto.getPlaceImg());\nplace.setPlaceInfo(dto.getPlaceInfo());\nPlace saved = placeRepository.save(place);\nreturn new PlaceResDto(saved);',""),
    ("updatePlace\n(PlaceServiceImpl)",
     'Place place = placeRepository.findById(placeId)\n    .orElseThrow(() -> new BoardException("없음"));\nplace.setPlaceName(dto.getPlaceName());\nplace.setPlaceImg(dto.getPlaceImg());\nplace.setPlaceInfo(dto.getPlaceInfo());\nreturn new PlaceResDto(place);',
     "@Transactional → save() 불필요"),
    ("deletePlace\n(PlaceServiceImpl)",
     'Place place = placeRepository.findById(placeId)\n    .orElseThrow(() -> new BoardException("없음"));\nplaceRepository.delete(place);',""),
    ("sendRequest\n(DateRequestServiceImpl)",
     'Board board = boardRepository.findById(boardId)\n    .orElseThrow(() -> new BoardException("없음"));\nif (board.getMember().getMemberNo().equals(senderNo)) {\n    throw new BoardException("본인 게시글 신청 불가");\n}\nMember sender = memberRepository.findById(senderNo)\n    .orElseThrow(() -> new MemberException("없음"));\nDateRequest req = new DateRequest();\nreq.setMessage(dto.getMessage());\nreq.setBoard(board);\nreq.setSender(sender);\nDateRequest saved = dateRequestRepository.save(req);\nreturn new DateRequestResDto(saved);',
     "status 기본값 PENDING"),
    ("getReceivedRequestsByBoard\n(DateRequestServiceImpl)",
     'Board board = boardRepository.findById(boardId)\n    .orElseThrow(() -> new BoardException("없음"));\nif (!board.getMember().getMemberNo().equals(memberNo)) {\n    throw new BoardException("본인 게시글만 조회 가능");\n}\nreturn dateRequestRepository.findByBoard_BoardId(boardId).stream()\n    .map(r -> new DateRequestResDto(r)).toList();',
     "게시글 작성자만"),
    ("getSentRequests\n(DateRequestServiceImpl)",
     'return dateRequestRepository.findBySender_MemberNo(memberNo).stream()\n    .map(r -> new DateRequestResDto(r)).toList();',""),
    ("acceptRequest\n(DateRequestServiceImpl)",
     'DateRequest req = dateRequestRepository.findById(requestId)\n    .orElseThrow(() -> new BoardException("없음"));\nif (!req.getBoard().getMember().getMemberNo().equals(memberNo)) {\n    throw new BoardException("수락 권한 없음");\n}\nreq.setStatus("ACCEPTED");\nreturn new DateRequestResDto(req);',
     "★ ACCEPTED → 매칭 성공"),
    ("rejectRequest\n(DateRequestServiceImpl)",
     'DateRequest req = dateRequestRepository.findById(requestId)\n    .orElseThrow(() -> new BoardException("없음"));\nif (!req.getBoard().getMember().getMemberNo().equals(memberNo)) {\n    throw new BoardException("거절 권한 없음");\n}\nreq.setStatus("REJECTED");\nreturn new DateRequestResDto(req);',
     "status → REJECTED"),
    ("cancelRequest\n(DateRequestServiceImpl)",
     'DateRequest req = dateRequestRepository.findById(requestId)\n    .orElseThrow(() -> new BoardException("없음"));\nif (!req.getSender().getMemberNo().equals(memberNo)) {\n    throw new BoardException("본인 신청만 취소 가능");\n}\ndateRequestRepository.delete(req);',""),
]

for ri,(name,code,desc) in enumerate(svc_items,3):
    bg=C_ODD if ri%2==1 else C_EVEN
    c1=ws4.cell(row=ri,column=1,value=name)
    c1.font=bf(bold=True,sz=9); c1.fill=fl(C_BLUE); c1.alignment=al(wrap=True); c1.border=bd()
    c2=ws4.cell(row=ri,column=2,value=code)
    c2.font=cf(8); c2.fill=fl(bg); c2.alignment=al(wrap=True); c2.border=bd()
    c3=ws4.cell(row=ri,column=3,value=desc)
    c3.font=bf(sz=9); c3.fill=fl(C_YELLOW if ri%2==1 else C_GREEN)
    c3.alignment=al(wrap=True); c3.border=bd()
    ws4.row_dimensions[ri].height=100

ws4.freeze_panes="A3"


# ══════════════════════════════════════════════════════════
# 시트 6 - 에러 처리 가이드
# ══════════════════════════════════════════════════════════
ws7=wb.create_sheet("에러 처리 가이드")
title(ws7,1,"에러 처리 - Exception 클래스 완성 코드","D")
cw(ws7,[6,20,54,26])
hrow(ws7,2,["구분","파일","완성 코드","설명"])

errs=[
    ("완성","MemberException.java",
     'import org.springframework.http.HttpStatus;\n\npublic class MemberException extends RuntimeException {\n\n    private final HttpStatus httpStatus;\n\n    public MemberException(String message) {\n        super(message);\n        this.httpStatus = HttpStatus.BAD_REQUEST;\n    }\n\n    public MemberException(String message, HttpStatus httpStatus) {\n        super(message);\n        this.httpStatus = httpStatus;\n    }\n\n    public HttpStatus getHttpStatus() {\n        return httpStatus;\n    }\n}',
     "MemberException.java에 그대로 작성"),
    ("완성","BoardException.java",
     "MemberException과 동일한 구조\n클래스명만 BoardException으로 변경",
     ""),
    ("완성","GlobalExceptionAdvice.java",
     '@RestControllerAdvice\npublic class GlobalExceptionAdvice {\n\n    @ExceptionHandler(MemberException.class)\n    public ResponseEntity<Map<String, String>> handleMemberException(MemberException e) {\n        Map<String, String> body = new HashMap<>();\n        body.put("message", e.getMessage());\n        body.put("status", String.valueOf(e.getHttpStatus().value()));\n        return ResponseEntity.status(e.getHttpStatus()).body(body);\n    }\n\n    @ExceptionHandler(BoardException.class)\n    public ResponseEntity<Map<String, String>> handleBoardException(BoardException e) {\n        Map<String, String> body = new HashMap<>();\n        body.put("message", e.getMessage());\n        body.put("status", String.valueOf(e.getHttpStatus().value()));\n        return ResponseEntity.status(e.getHttpStatus()).body(body);\n    }\n}',
     "GlobalExceptionAdvice.java\n완성 코드"),
    ("400","Bad Request",
     'throw new MemberException("이미 사용 중인 ID입니다.");\nthrow new BoardException("본인 게시글에는 신청할 수 없습니다.");',
     "기본 400 (BAD_REQUEST)"),
    ("404","Not Found",
     'throw new MemberException("회원 없음", HttpStatus.NOT_FOUND);\nthrow new BoardException("게시글 없음", HttpStatus.NOT_FOUND);',
     "데이터 없음 (NOT_FOUND)"),
    ("200","OK","return ResponseEntity.ok(result);","조회/수정 성공"),
    ("201","Created","return ResponseEntity.status(HttpStatus.CREATED).body(result);","등록 성공"),
    ("204","No Content","return ResponseEntity.noContent().build();","삭제 성공"),
]

code_bg={"완성":"D1ECF1","400":C_DEL,"404":C_PUT,"200":C_GET,"201":C_POST,"204":C_DEL}
for ri,(code,name,exc,when) in enumerate(errs,3):
    bg=C_ODD if ri%2==1 else C_EVEN
    c1=ws7.cell(row=ri,column=1,value=code)
    c1.font=Font(name="맑은 고딕",bold=True,size=9)
    c1.fill=fl(code_bg.get(code,bg)); c1.alignment=al("center"); c1.border=bd()
    wc(ws7,ri,2,name,bold=True,bg=bg,halign="center")
    c3=ws7.cell(row=ri,column=3,value=exc)
    c3.font=cf(8); c3.fill=fl(bg); c3.alignment=al(wrap=True); c3.border=bd()
    wc(ws7,ri,4,when,bg=bg,wrap=True)
    ws7.row_dimensions[ri].height=130 if code=="완성" else 35


# ══════════════════════════════════════════════════════════
# 저장
# ══════════════════════════════════════════════════════════
path="/Users/hongjunhwa/Desktop/Personal_Project/work/lov3_board_project/LOV3_게시판_REST_API_명세서.xlsx"
wb.save(path)
print(f"완료: {path}")
